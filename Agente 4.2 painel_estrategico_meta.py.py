import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from langchain.llms import Ollama
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain

# === Caminhos ===
base_path = r"C:\Users\britt\Downloads\Projeto_Final_IA2A"
graficos_dir = os.path.join(base_path, "output", "graficos")
os.makedirs(graficos_dir, exist_ok=True)

# === IA Gemma via LangChain ===
llm = Ollama(model="gemma3:4b")
template = PromptTemplate(
    input_variables=["indicador", "valor"],
    template="Explique de forma estratégica e clara o que significa o indicador '{indicador}' com valor de {valor}% para o desempenho de vendas da empresa."
)
chain = LLMChain(llm=llm, prompt=template)

# === Função para extrair percentuais de texto ===
def extrair_percentuais(texto):
    padroes = {
        "Faturamento": r"faturamento.*?(\d+\.?\d*)%",
        "Margem_Produtos": r"margem.*?(\d+\.?\d*)%",
        "Volume_Vendas": r"volume.*?(\d+\.?\d*)%",
        "Crescimento_Vendas": r"crescimento.*?vendas.*?(\d+\.?\d*)%",
        "Rentabilidade": r"rentabilidade.*?(\d+\.?\d*)%",
    }
    resultados = {}
    for nome, padrao in padroes.items():
        match = re.search(padrao, texto, re.IGNORECASE)
        if match:
            resultados[nome] = float(match.group(1))
    return resultados

# === Função para gerar gráfico de barras com explicações da IA ===
def gerar_grafico_barras(dados, titulo, nome_arquivo):
    plt.figure(figsize=(8, 5))
    nomes = list(dados.keys())
    valores = list(dados.values())
    cores = ["blue", "purple", "green", "orange", "teal"]
    plt.bar(nomes, valores, color=cores[:len(nomes)])
    plt.title(titulo)
    plt.ylabel("Percentual (%)")
    for i, valor in enumerate(valores):
        explicacao = chain.run({"indicador": nomes[i], "valor": valor})
        plt.text(i, valor + 1, f"{valor:.1f}%\n{explicacao[:80]}...", ha='center', fontsize=8)
    plt.grid(axis="y", linestyle="--", alpha=0.7)
    plt.tight_layout()
    caminho_completo = os.path.join(graficos_dir, nome_arquivo)
    plt.savefig(caminho_completo)
    plt.close()
    print(f"✅ Gráfico salvo em: {caminho_completo}")

# === Função para gerar gráfico de linha com explicação da IA ===
def gerar_grafico_linha(df, coluna, titulo, nome_arquivo, ylabel):
    if "Ano" in df.columns and coluna in df.columns:
        plt.figure(figsize=(8, 5))
        plt.plot(df["Ano"], df[coluna], marker='o', color='orange')
        plt.title(titulo)
        plt.xlabel("Ano")
        plt.ylabel(ylabel)
        for i in range(len(df)):
            valor = df[coluna][i]
            explicacao = chain.run({"indicador": coluna, "valor": valor})
            plt.text(df["Ano"][i], valor, f"{valor:.1f}%\n{explicacao[:80]}...", fontsize=8)
        plt.grid(True)
        plt.tight_layout()
        caminho_completo = os.path.join(graficos_dir, nome_arquivo)
        plt.savefig(caminho_completo)
        plt.close()
        print(f"✅ Gráfico salvo em: {caminho_completo}")

# === Varredura de arquivos ===
dados_extraidos = {}

for arquivo in os.listdir(base_path):
    caminho = os.path.join(base_path, arquivo)

    # TXT
    if arquivo.endswith(".txt"):
        with open(caminho, "r", encoding="utf-8") as f:
            texto = f.read()
            dados = extrair_percentuais(texto)
            dados_extraidos.update(dados)

    # CSV
    elif arquivo.endswith(".csv"):
        try:
            df = pd.read_csv(caminho)
            for coluna in df.columns:
                if coluna != "Ano" and df[coluna].dtype in ["float64", "int64"]:
                    gerar_grafico_linha(df, coluna, f"{coluna} ao longo dos anos", f"grafico_{coluna.lower()}.png", f"{coluna} (%)")
        except Exception as e:
            print(f"⚠️ Erro ao ler CSV {arquivo}: {e}")

    # XLSX
    elif arquivo.endswith(".xlsx"):
        try:
            wb = load_workbook(caminho)
            for sheet in wb.sheetnames:
                df = pd.DataFrame(wb[sheet].values)
                df.columns = df.iloc[0]
                df = df[1:]
                if "Ano" in df.columns:
                    for coluna in df.columns:
                        if coluna != "Ano":
                            try:
                                df[coluna] = pd.to_numeric(df[coluna])
                                gerar_grafico_linha(df, coluna, f"{coluna} ao longo dos anos", f"grafico_{coluna.lower()}.png", f"{coluna} (%)")
                            except:
                                continue
        except Exception as e:
            print(f"⚠️ Erro ao ler XLSX {arquivo}: {e}")

# === Gráfico de indicadores extraídos dos textos ===
if dados_extraidos:
    gerar_grafico_barras(dados_extraidos, "📦 Indicadores Estratégicos de Vendas", "grafico_indicadores_vendas.png")
else:
    print("⚠️ Nenhum dado estratégico extraído dos arquivos .txt")
